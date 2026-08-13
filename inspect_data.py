import os
import openpyxl
import pandas as pd

def inspect_doc(filename):
    print("=== INSPECTING", filename, "===")
    with open(filename, 'rb') as f:
        head = f.read(100)
        print("Header bytes:", head[:30])
        # check if it's OLE CFB or docx (zip) or xml/rtf
        if head.startswith(b'PK'):
            print("Format: DOCX (Zip archive)")
        elif b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1' in head:
            print("Format: Word 97-2003 Binary (.doc)")
        elif b'{\\rtf' in head:
            print("Format: RTF")
        elif b'<?xml' in head or b'<html' in head or b'<HTML' in head:
            print("Format: XML/HTML")
        else:
            print("Unknown header format")

def inspect_xlsx(filename):
    print("\n=== INSPECTING EXCEL", filename, "===")
    wb = openpyxl.load_workbook(filename, data_only=True)
    print("Sheet names:", wb.sheetnames)
    for sheetname in wb.sheetnames[:2]:
        ws = wb[sheetname]
        print(f"\n--- Sheet: {sheetname} (max_row={ws.max_row}, max_col={ws.max_column}) ---")
        for r in range(1, min(15, ws.max_row + 1)):
            row_vals = [ws.cell(row=r, column=c).value for c in range(1, min(15, ws.max_column + 1))]
            if any(v is not None for v in row_vals):
                print(f"Row {r}: {row_vals}")

inspect_doc("Advanced learners template 25-26.doc")
inspect_doc("Slow learners template 25-26.doc")
inspect_xlsx("CSE-C.xlsx")
inspect_xlsx("sem ii results for 24-25 2nd year use.xlsx")
