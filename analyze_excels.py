import openpyxl
import pandas as pd

def analyze_excel(filename):
    print(f"\n==================== ANALYZING {filename} ====================")
    wb = openpyxl.load_workbook(filename, data_only=True)
    for sname in wb.sheetnames:
        ws = wb[sname]
        print(f"\n--- Sheet: {sname} ---")
        rows = []
        for r in range(1, min(20, ws.max_row + 1)):
            row_data = [ws.cell(row=r, column=c).value for c in range(1, min(25, ws.max_column + 1))]
            if any(v is not None for v in row_data):
                rows.append((r, row_data))
        for r_idx, r_data in rows[:10]:
            print(f"Row {r_idx:2d}: {r_data}")

analyze_excel("CSE-C.xlsx")
analyze_excel("sem ii results for 24-25 2nd year use.xlsx")
